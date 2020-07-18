
from openpyxl import load_workbook, Workbook
import os

def excel_to_dict(file):
    wb = load_workbook(file)
    wb_dict = {}

    for sheet in wb.worksheets:
        wb_dict[sheet.title] = {}

        for row in sheet.iter_rows(min_row=3, min_col=1):
            wb_dict[sheet.title][row[0].value] = {}

            for cell in row:
                k = sheet.cell(row=2, column=cell.column).value
                val = cell.value
                wb_dict[sheet.title][row[0].value][k] = val
    return wb_dict

def write_invoice(file, invoice_no):
    wb_dict = excel_to_dict(file)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # head
    ws['A1'] = 'Invoice #'
    ws['A2'] = 'Date'
    ws['C1'] = 'Customer'
    ws['C2'] = 'Phone'
    ws['C3'] = 'Address'
    ws['B1'] = invoice_no
    ws['B2'] = wb_dict['Invoices'][invoice_no]['invoice_date']
    customer_id = wb_dict['Invoices'][invoice_no]['customer_id']
    id_info = wb_dict['Customers'][customer_id]
    ws['D1'] = id_info['f_name'] + ' ' + id_info['l_name']
    ws['D2'] = id_info['phone']
    ws['D3'] = f"{id_info['street']} {id_info['city']}"
 
    # items
    ws['A5'] = 'Items'
    ws['B5'] = 'Unit price'
    ws['C5'] = 'Quantity'
    ws['D5'] = 'Price'
    item_info = []
    for key in wb_dict['Invoice Line Items']:
        item = wb_dict['Invoice Line Items'][key]
        if item['invoice_no'] == invoice_no:
            product_id = item['product_id']
            item['name'] = wb_dict['Product'][product_id]['name']
            item['unit_price'] = wb_dict['Product'][product_id]['unit_price']
            item_info.append(item)
    i = 0
    subtotal = 0
    for item in item_info:
        ws[f'A{6+i}'] = item['name']
        ws[f'B{6+i}'] = item['unit_price']
        ws[f'C{6+i}'] = item['quantity']
        ws[f'D{6+i}'] = item['unit_price'] * item['quantity']
        i += 1
        subtotal += item['unit_price'] * item['quantity']

    # total
    ws[f'C{7+len(item_info)}'] = 'Subtotal'
    ws[f'C{8+len(item_info)}'] = 'GST'
    ws[f'C{9+len(item_info)}'] = 'Total'
    ws[f'D{7+len(item_info)}'] = subtotal
    ws[f'D{8+len(item_info)}'] = subtotal * 0.05
    ws[f'D{9+len(item_info)}'] = subtotal * 1.05


    wb.save(os.getcwd() + f'/Invoice_{invoice_no}.xlsx')


file = os.getcwd() + '/comp_230.xlsx'

try:
    wb_dict = excel_to_dict(file)
except:
    print(f'**** Fail to load file ****')
    raise


while True:   
    invoice_no = int(input('Enter the invoice number to print: '))

    if not invoice_no in wb_dict['Invoices']:
        print(f'Invoice # {invoice_no} not found')
        continue
    else:
        break

try:
    write_invoice(file, invoice_no)
    print(f'Invoice_{invoice_no}.xlsx has been created successfully')
except:
    print('Failed to create invoice')
    raise